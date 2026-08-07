"""FEHD list sync for seating.xlsx.

Run by GitHub Actions on a schedule, or on demand via 'Check FEHD.command'.
Steps, in order:
1. Read the "List updated on ..." date from the FEHD PDF and compare it
   with cell B1 of the Meta sheet. If they match, stop — nothing is
   downloaded and nothing is changed.
2. Otherwise fetch the live list, and back up the current sheet as
   'seating prior.xlsx' BEFORE changing anything.
3. Move rows for delisted licences to the 'Delisted' sheet (seating value
   preserved, so a relisted restaurant can be restored).
4. Append rows for restaurants newly added by FEHD (seating left blank).
5. Refresh the Meta sheet, write summary.txt for the commit message, and
   refresh docs/status.json for the phone status page.

Every exit path writes docs/status.json, so the phone page can always show
when the list was last checked — not just when it last changed.
Pass --force to sync even when the dates match.
"""
import datetime
import json
import os
import re
import shutil
import sys

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pypdf import PdfReader

SHEET = "seating.xlsx"
BACKUP = "seating prior.xlsx"
STATUS = os.path.join("docs", "status.json")
DATA_URL = "https://www.fehd.gov.hk/english/licensing/dog_restaurants/getData.php"
PDF_URL = "https://www.fehd.gov.hk/english/licensing/dog_restaurants/fulllist.pdf"
REFERER = "https://www.fehd.gov.hk/english/licensing/dog_restaurants/dog_restaurants_list.html"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
HEADERS = ["FEHD Licence", "Name (English)", "Name (中文)", "District", "Address", "Seating"]
MIN_RECORDS = 500
HISTORY_KEPT = 20

force = "--force" in sys.argv
HKT = datetime.timezone(datetime.timedelta(hours=8))


def now_hkt():
    """Timestamp for the Meta sheet. CI runs in UTC, so pin the zone."""
    return datetime.datetime.now(HKT).strftime("%Y-%m-%d %H:%M")


def write_status(state, fehd_date=None, total=None, change=None, message=""):
    """Refresh docs/status.json. Called on every exit path, including errors."""
    previous = {}
    if os.path.exists(STATUS):
        try:
            with open(STATUS, encoding="utf-8") as f:
                previous = json.load(f)
        except (ValueError, OSError):
            previous = {}

    payload = {
        "state": state,
        "message": message,
        "last_checked_hkt": now_hkt(),
        "fehd_date": fehd_date or previous.get("fehd_date"),
        "total": total if total is not None else previous.get("total"),
        "last_change": change or previous.get("last_change"),
        "history": previous.get("history", []),
    }
    if change:
        payload["history"] = ([{
            "checked": payload["last_checked_hkt"],
            "fehd_date": change["fehd_date"],
            "added": len(change["added"]),
            "removed": len(change["removed"]),
        }] + payload["history"])[:HISTORY_KEPT]

    os.makedirs(os.path.dirname(STATUS), exist_ok=True)
    with open(STATUS, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")


def fail(msg):
    """Abort loudly. The sheet is never left half-written at this point."""
    write_status("error", message=msg)
    sys.exit(f"ERROR: {msg}")


def normalise_date(text):
    """'Aug 6' / 'August 6' -> ('August 6', month, day).

    FEHD prints no year, so only month and day are compared.
    """
    if not text:
        return None
    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    lookup = {m.lower(): i for i, m in enumerate(months, 1)}
    lookup.update({m[:3].lower(): i for i, m in enumerate(months, 1)})
    t = str(text).strip().replace(",", " ")
    name = re.search(r"[A-Za-z]{3,}", t)
    day = re.search(r"\b(\d{1,2})\b", t)
    if not name or not day:
        return None
    month = lookup.get(name.group(0).lower())
    if month is None:
        return None
    return f"{months[month - 1]} {int(day.group(1))}", month, int(day.group(1))


# --- 1. Compare the FEHD publication date against the Meta sheet ----------
wb = load_workbook(SHEET)
if "Seating" not in wb.sheetnames:
    fail(f"'{SHEET}' has no 'Seating' sheet")
ws = wb["Seating"]
meta = wb["Meta"] if "Meta" in wb.sheetnames else None

try:
    resp = requests.get(PDF_URL, headers={"User-Agent": UA, "Referer": REFERER}, timeout=60)
    resp.raise_for_status()
except requests.RequestException as e:
    fail(f"could not reach the FEHD website: {e}")

with open("fulllist.pdf", "wb") as f:
    f.write(resp.content)
page1 = PdfReader("fulllist.pdf").pages[0].extract_text() or ""
found = re.search(r"List updated on\s+([A-Za-z]+\.?\s*\d{1,2})", page1)
if not found:
    fail("could not find 'List updated on ...' in the FEHD PDF — the page layout may have changed")
fehd_date = normalise_date(found.group(1))
if not fehd_date:
    fail(f"could not read the FEHD date from {found.group(1)!r}")

known_date = normalise_date(meta["B1"].value) if meta else None
sheet_total = sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
                  if r[0] is not None and str(r[0]).strip())

if known_date and fehd_date[1:] == known_date[1:] and not force:
    write_status("up_to_date", fehd_date[0], sheet_total)
    print(f"No update required — FEHD list still dated {fehd_date[0]}.")
    sys.exit(0)

# --- 2. Fetch the live list and back up before touching anything ---------
try:
    resp = requests.get(DATA_URL, headers={
        "Accept": "application/json", "User-Agent": UA, "Referer": REFERER,
    }, timeout=60)
    resp.raise_for_status()
    data = resp.json()
except (requests.RequestException, ValueError) as e:
    fail(f"could not download the FEHD restaurant list: {e}")

records = data if isinstance(data, list) else next(iter(data.values()))
live = {}
for x in records:
    lic = str(x.get("licence", "")).strip()
    if lic:
        live[lic] = x
if len(live) < MIN_RECORDS:
    fail(f"FEHD returned only {len(live)} records — refusing to proceed")

rows = {}
for r in range(2, ws.max_row + 1):
    lic = ws.cell(row=r, column=1).value
    if lic is not None and str(lic).strip():
        rows[str(lic).strip()] = r

added = sorted(set(live) - set(rows), key=lambda l: (
    live[l].get("district_en", ""), live[l].get("shop_sign_en", "")))
removed = sorted(set(rows) - set(live))

if not added and not removed:
    # Date moved but the roster is identical — just record the new date.
    if meta:
        meta["B1"] = fehd_date[0]
        meta["B4"] = now_hkt()
        wb.save(SHEET)
    write_status("up_to_date", fehd_date[0], sheet_total,
                 message=f"FEHD re-dated to {fehd_date[0]}; no restaurants changed.")
    with open("summary.txt", "w", encoding="utf-8") as f:
        f.write(f"FEHD re-dated to {fehd_date[0]} (no restaurant changes)\n")
    print(f"FEHD re-dated to {fehd_date[0]}, but no restaurants changed.")
    sys.exit(0)

shutil.copyfile(SHEET, BACKUP)

# --- 3. Move delisted rows to the 'Delisted' sheet -----------------------
delisted_detail = []
if removed:
    if "Delisted" in wb.sheetnames:
        dws = wb["Delisted"]
    else:
        dws = wb.create_sheet("Delisted")
        for c, h in enumerate(HEADERS + ["Delisted on"], 1):
            cell = dws.cell(row=1, column=c, value=h)
            cell.font = Font(name="Arial", size=10, bold=True)
        for col, width in zip("ABCDEFG", [14, 34, 26, 14, 52, 12, 14]):
            dws.column_dimensions[col].width = width
        dws.freeze_panes = "A2"

    today = datetime.datetime.now(HKT).date().isoformat()
    for lic in removed:
        r = rows[lic]
        values = [ws.cell(row=r, column=c).value for c in range(1, 7)]
        delisted_detail.append((lic, values[1], values[2], values[5]))
        dr = dws.max_row + 1
        for c, v in enumerate(values + [today], 1):
            cell = dws.cell(row=dr, column=c, value=v)
            cell.font = Font(name="Arial", size=10)
            if c == 1:
                cell.number_format = "@"

    # Delete bottom-up so earlier row numbers stay valid.
    for r in sorted((rows[l] for l in removed), reverse=True):
        ws.delete_rows(r, 1)

# --- 4. Append the new restaurants ---------------------------------------
yellow = PatternFill("solid", fgColor="FFF3C4")
for x in added:
    rec = live[x]
    r = ws.max_row + 1
    values = [str(rec.get("licence", "")).strip(), rec.get("shop_sign_en", ""),
              rec.get("shop_sign_tc", ""), rec.get("district_en", ""),
              rec.get("address_en", ""), ""]
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Arial", size=10)
        if c == 1:
            cell.number_format = "@"
        if c == 6:
            cell.fill = yellow
            cell.font = Font(name="Arial", size=10, bold=True)

# Rebuild the dropdown and filter over the new extent.
last = ws.max_row
ws.data_validations.dataValidation = []
dv = DataValidation(type="list", formula1='"indoor,outdoor"', allow_blank=True)
ws.add_data_validation(dv)
dv.sqref = f"F2:F{last}"
ws.auto_filter.ref = f"A1:F{last}"

# --- 5. Refresh Meta, the commit summary, and the phone status ----------
if meta:
    meta["B1"] = fehd_date[0]
    meta["B3"] = last - 1
    meta["B4"] = now_hkt()

wb.save(SHEET)

change = {
    "date": datetime.datetime.now(HKT).date().isoformat(),
    "fehd_date": fehd_date[0],
    "added": [{"licence": x, "en": live[x].get("shop_sign_en", ""),
               "tc": live[x].get("shop_sign_tc", ""),
               "district": live[x].get("district_en", "")} for x in added],
    "removed": [{"licence": l, "en": en, "tc": tc, "seating": seat or ""}
                for l, en, tc, seat in delisted_detail],
}
write_status("updated", fehd_date[0], last - 1, change=change)

lines = [f"FEHD sync {fehd_date[0]}: {len(added)} added, {len(removed)} delisted", ""]
for x in added:
    rec = live[x]
    lines.append(f"+ [{x}] {rec.get('shop_sign_en', '')} / "
                 f"{rec.get('shop_sign_tc', '')} ({rec.get('district_en', '')})")
for lic, en, tc, seat in delisted_detail:
    kept = f" (seating '{seat}' preserved)" if seat else ""
    lines.append(f"- [{lic}] {en} / {tc} → moved to Delisted sheet{kept}")
with open("summary.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(lines) + "\n")

# Machine-readable copy for the Mac popup.
with open("summary.json", "w", encoding="utf-8") as f:
    json.dump(change, f, ensure_ascii=False, indent=2)

print("\n".join(lines))
