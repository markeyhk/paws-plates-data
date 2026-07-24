"""Regenerates seating.json from seating.xlsx.

Run automatically by GitHub Actions whenever seating.xlsx changes.
Reads the 'Seating' sheet: column A = FEHD licence, column F = seating
value ('indoor' / 'outdoor' / blank). Blank rows produce no entry.
"""
import datetime
import json
import sys

from openpyxl import load_workbook

wb = load_workbook("seating.xlsx", data_only=True)
ws = wb["Seating"]

arrangements = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    licence, seating = row[0], row[5]
    if licence is None or seating is None:
        continue
    licence = str(licence).strip()
    value = str(seating).strip().lower()
    if not licence:
        continue
    if value in ("indoor", "outdoor"):
        arrangements[licence] = value
    elif value:
        print(f"WARNING: ignoring unrecognised seating value '{seating}' for licence {licence}")

if not arrangements:
    sys.exit("Refusing to write an empty seating.json — check seating.xlsx")

payload = {
    "updated": datetime.date.today().isoformat(),
    "arrangements": arrangements,
}
with open("seating.json", "w", encoding="utf-8") as f:
    json.dump(payload, f, ensure_ascii=False, indent=2)
    f.write("\n")

print(f"seating.json written with {len(arrangements)} entries")
